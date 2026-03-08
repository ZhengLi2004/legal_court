from __future__ import annotations

import argparse
import json
from html import escape
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation


def _load_jsonl(path: Path) -> list[dict[str, Any]]:
    return [
        json.loads(line)
        for line in path.read_text(encoding="utf-8").splitlines()
        if line.strip()
    ]


def _write_text(path: Path, text: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(text, encoding="utf-8")


def _slug(text: str) -> str:
    return "".join(ch if ch.isalnum() or ch in ("_", "-") else "_" for ch in text)


BASE_CSS = """
body { font-family: 'Noto Sans SC', 'Microsoft YaHei', sans-serif; margin: 0; background: #f6f5f1; color: #1f2937; }
.page { max-width: 1080px; margin: 0 auto; padding: 32px 28px 80px; }
.header { background: linear-gradient(135deg, #1f3a5f, #365f8f); color: white; padding: 24px 28px; border-radius: 16px; margin-bottom: 24px; }
.header h1 { margin: 0 0 8px; font-size: 28px; }
.header p { margin: 0; opacity: 0.9; }
.meta { display: flex; gap: 12px; flex-wrap: wrap; margin-top: 14px; }
.badge { background: rgba(255,255,255,0.18); border: 1px solid rgba(255,255,255,0.22); padding: 6px 10px; border-radius: 999px; font-size: 14px; }
.panel { background: white; border-radius: 16px; padding: 22px 24px; margin-bottom: 18px; box-shadow: 0 8px 24px rgba(15, 23, 42, 0.06); }
.panel h2 { margin: 0 0 12px; font-size: 20px; }
.panel.claim { border-left: 8px solid #6d8b74; }
.panel.plaintiff { border-left: 8px solid #d3a738; }
.panel.judgment { border-left: 8px solid #457b9d; }
.panel.opinion { border-left: 8px solid #5b8c5a; }
.panel.tip { border-left: 8px solid #8a6f5a; }
.prewrap { white-space: pre-wrap; line-height: 1.8; font-size: 15px; }
.nav { display: flex; justify-content: space-between; gap: 12px; margin-top: 18px; }
.nav a, .index-table a { color: #1d4ed8; text-decoration: none; }
.index-table { width: 100%; border-collapse: collapse; background: white; border-radius: 14px; overflow: hidden; box-shadow: 0 8px 24px rgba(15, 23, 42, 0.06); }
.index-table th, .index-table td { padding: 12px 14px; border-bottom: 1px solid #e5e7eb; text-align: left; vertical-align: top; }
.index-table th { background: #ecf2f9; }
.note { color: #6b7280; font-size: 14px; }
code { background: #f3f4f6; padding: 2px 6px; border-radius: 6px; }
""".strip()


def _render_claim_card(
    row: dict[str, Any], prev_name: str | None, next_name: str | None
) -> str:
    nav_left = (
        f'<a href="{escape(prev_name)}">上一份</a>' if prev_name else "<span></span>"
    )

    nav_right = (
        f'<a href="{escape(next_name)}">下一份</a>' if next_name else "<span></span>"
    )

    plaintiff_text = escape(str(row.get("plaintiff_text", "") or ""))

    return f"""<!doctype html>
<html lang="zh-CN">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>{escape(str(row["anchor_case_id"]))} - Claim 复核</title>
  <style>{BASE_CSS}</style>
</head>
<body>
  <div class="page">
    <div class="header">
      <h1>Claim 复核卡片</h1>
      <p>只阅读原告诉称，并在工作簿中精确复制实体诉求原文。</p>
      <div class="meta">
        <span class="badge">{escape(str(row["anchor_case_id"]))}</span>
        <span class="badge">阶段：{escape(str(row.get("stage", "")))}</span>
      </div>
    </div>
    <div class="panel tip">
      <h2>填写规则</h2>
      <div class="prewrap">1. 只填写实体诉求。
2. 必须从原文精确复制，不要改写，不要摘要。
3. 诉讼费等程序性请求不要写入。
4. 请回到 Excel 工作簿的 Claim 答案表填写。</div>
    </div>
    <div class="panel plaintiff">
      <h2>原告诉称</h2>
      <div class="prewrap">{plaintiff_text}</div>
    </div>
    <div class="nav">
      <div>{nav_left}</div>
      <div><a href="index.html">返回目录</a></div>
      <div>{nav_right}</div>
    </div>
  </div>
</body>
</html>"""


def _render_status_card(
    row: dict[str, Any], prev_name: str | None, next_name: str | None
) -> str:
    nav_left = (
        f'<a href="{escape(prev_name)}">上一份</a>' if prev_name else "<span></span>"
    )

    nav_right = (
        f'<a href="{escape(next_name)}">下一份</a>' if next_name else "<span></span>"
    )

    claim_text = escape(
        str((row.get("claims_review") or [{}])[0].get("claim_text_raw", "") or "")
    )

    plaintiff_text = escape(str(row.get("plaintiff_text", "") or ""))
    judgment_text = escape(str(row.get("judgment_result_text", "") or ""))
    opinion_text = escape(str(row.get("court_opinion_text", "") or ""))

    return f"""<!doctype html>
<html lang="zh-CN">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>{escape(str(row["anchor_case_id"]))} - Status 复核</title>
  <style>{BASE_CSS}</style>
</head>
<body>
  <div class="page">
    <div class="header">
      <h1>Status 复核卡片</h1>
      <p>阅读 claim、裁判结果和法院观点，并在工作簿中选择三态标签。</p>
      <div class="meta">
        <span class="badge">{escape(str(row["anchor_case_id"]))}</span>
        <span class="badge">阶段：{escape(str(row.get("stage", "")))}</span>
      </div>
    </div>
    <div class="panel tip">
      <h2>填写规则</h2>
      <div class="prewrap">1. 不要改动 claim 文本。
2. 只填写 `ACCEPTED / REJECTED / UNMENTIONED`。
3. 仅部分支持、金额调减、范围缩窄统一记为 `UNMENTIONED`。
4. 请回到 Excel 工作簿的 Status 答案表填写。</div>
    </div>
    <div class="panel claim">
      <h2>待判定 Claim</h2>
      <div class="prewrap">{claim_text}</div>
    </div>
    <div class="panel plaintiff">
      <h2>原告诉称</h2>
      <div class="prewrap">{plaintiff_text}</div>
    </div>
    <div class="panel judgment">
      <h2>裁判结果</h2>
      <div class="prewrap">{judgment_text}</div>
    </div>
    <div class="panel opinion">
      <h2>法院观点</h2>
      <div class="prewrap">{opinion_text}</div>
    </div>
    <div class="nav">
      <div>{nav_left}</div>
      <div><a href="index.html">返回目录</a></div>
      <div>{nav_right}</div>
    </div>
  </div>
</body>
</html>"""


def _render_index(
    title: str, subtitle: str, rows: list[dict[str, Any]], card_dir_name: str, task: str
) -> str:
    body_rows = []

    for row in rows:
        anchor = str(row["anchor_case_id"])
        stage = str(row.get("stage", "") or "")
        card_name = f"{_slug(anchor)}.html"

        if task == "claim":
            preview = escape(str(row.get("plaintiff_text", "") or "")[:80]).replace(
                "\n", " "
            )

        else:
            preview = escape(
                str(
                    (row.get("claims_review") or [{}])[0].get("claim_text_raw", "")
                    or ""
                )
            )

        body_rows.append(
            f'<tr><td>{escape(anchor)}</td><td>{escape(stage)}</td><td>{preview}</td><td><a href="{escape(card_dir_name)}/{escape(card_name)}">打开</a></td></tr>'
        )

    rows_html = "".join(body_rows)

    return f"""<!doctype html>
<html lang="zh-CN">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>{escape(title)}</title>
  <style>{BASE_CSS}</style>
</head>
<body>
  <div class="page">
    <div class="header">
      <h1>{escape(title)}</h1>
      <p>{escape(subtitle)}</p>
    </div>
    <table class="index-table">
      <thead><tr><th>编号</th><th>阶段</th><th>预览</th><th>卡片</th></tr></thead>
      <tbody>{rows_html}</tbody>
    </table>
    <p class="note">建议先打开 Excel 工作簿，再通过“查看卡片”或本页链接阅读详细内容。</p>
  </div>
</body>
</html>"""


def _claim_guide_text() -> str:
    return """# Claim 复核细则

## 任务目标

你的任务不是“总结案件”，而是从 `原告诉称` 中找出应当进入数据集的**实体诉求**。

请把你认为应保留的诉求，原文精确复制到 `claim_review.xlsx` 的 `claim_1`、`claim_2` ... 列中。

## 什么算实体诉求

通常包括：

1. 支付、返还、赔偿
2. 确认、解除、撤销、认定无效
3. 继续履行、交付、过户、办理、停止侵权、排除妨害、恢复原状
4. 连带责任、优先受偿、抚养、迁葬等可独立裁判的法律效果

## 什么不要写进数据集

以下内容一般不算本任务中的 Claim：

1. 诉讼费、保全费、公告费、鉴定费等程序性费用承担
2. `撤销原判`、`发回重审`、`改判支持全部诉讼请求` 这类上诉外壳
3. 证据目录、证据名称、证人证言、情况说明
4. 说理性文字，如 `综上`、`事实认定错误`、`程序违法`
5. 从属细则，如道歉版式、道歉刊登天数、计算尾款的说明尾句

## 如何切分

规则是：**能否作为一个独立裁判单元**。

1. 如果一句里有两个独立救济效果，请拆开写。
2. 如果一句里只有一个核心诉求，后面只是金额计算方式或期限说明，不要把尾句单独再写一条。
3. 如果是 `撤销原判，并改判被告返还 10000 元`，只保留后面的实体尾巴：`被告返还10000元`。

## 必须遵守的格式规则

1. 必须从原文精确复制。
2. 不要改写，不要自己归纳。
3. 不要补充原文里没有出现的内容。
4. 如果你认为该案没有应保留的实体诉求，就把所有 `claim_*` 留空，并在 `notes` 中写 `无实体诉求`。

## 例子

### 例 1：应保留

原文：

`判令被告支付货款10000元，并承担违约金3000元。`

应填写：

- `claim_1 = 判令被告支付货款10000元`
- `claim_2 = 承担违约金3000元`

### 例 2：程序性费用不保留

原文：

`判令被告支付货款10000元，本案诉讼费由被告承担。`

应填写：

- `claim_1 = 判令被告支付货款10000元`

不要填写：

- `本案诉讼费由被告承担`

### 例 3：上诉外壳要去掉

原文：

`撤销原判，改判被上诉人返还借款50000元。`

应填写：

- `claim_1 = 被上诉人返还借款50000元`

不要填写：

- `撤销原判`

### 例 4：证据目录不保留

原文：

`为证明其主张，提交道路交通事故认定书一份。`

应填写：

- 留空

### 例 5：没有实体诉求

原文：

`请求驳回对方全部诉讼请求。`

应填写：

- 留空，并在 `notes` 里写 `无实体诉求`
"""


def _status_guide_text() -> str:
    return """# Status 复核细则

## 任务目标

你的任务是判断：给定的一条 Claim，在判决中被如何处理。

你只需要在 `status_review.xlsx` 中选择：

- `ACCEPTED`
- `REJECTED`
- `UNMENTIONED`

## 三个标签怎么理解

### ACCEPTED

法院对该 Claim 作出**明确支持**。

典型表现：

- `被告于本判决生效之日起十日内支付原告...`
- `确认合同无效`
- `解除双方签订的...合同`

### REJECTED

法院对该 Claim 作出**明确驳回**。

典型表现：

- `驳回原告该项诉讼请求`
- `驳回原告的全部/其他诉讼请求`
- 二审中：`驳回上诉，维持原判`，且该 Claim 明显属于上诉人的实体请求

### UNMENTIONED

以下情况统一记为 `UNMENTIONED`：

1. 判决没有明确处理该 Claim
2. 只支持了一部分，金额被调减
3. 范围被缩窄
4. 只支持了大 Claim 里的一个子项

## 判定顺序

1. 先看 `待判定 Claim`
2. 再看 `裁判结果`
3. 如果 `裁判结果` 不够清楚，再看 `法院观点`
4. 如果仍然不能明确判断为“完全支持”或“明确驳回”，就选 `UNMENTIONED`

## 必须遵守的格式规则

1. 不要修改 `claim_text_raw`
2. 只在 `status` 列选择三种标签之一
3. 如果拿不准，先选你认为最接近的标签，再把理由写到 `notes`

## 例子

### 例 1：明确支持

Claim：

`支付货款10000元`

裁判结果：

`被告于本判决生效之日起十日内支付原告货款10000元。`

应选：

- `ACCEPTED`

### 例 2：明确驳回

Claim：

`支付违约金5000元`

裁判结果：

`驳回原告关于违约金的诉讼请求。`

应选：

- `REJECTED`

### 例 3：部分支持

Claim：

`支付货款10000元`

裁判结果：

`被告支付原告货款6000元，驳回其余诉讼请求。`

应选：

- `UNMENTIONED`

原因：

这不是“完全支持”，也不是“整条明确驳回”，而是部分支持。

### 例 4：未明确处理

Claim：

`解除合同`

裁判结果只写：

`被告返还原告定金20000元。`

应选：

- `UNMENTIONED`

### 例 5：二审驳回上诉

Claim：

`返还借款50000元`

裁判结果：

`驳回上诉，维持原判。`

若该 Claim 显然是上诉人的实体上诉请求，应选：

- `REJECTED`
"""


def _apply_header_style(ws, header_row: int = 1) -> None:
    fill = PatternFill(fill_type="solid", fgColor="1F3A5F")
    font = Font(color="FFFFFF", bold=True)

    for cell in ws[header_row]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(vertical="top", wrap_text=True)


def _build_claim_workbook(rows: list[dict[str, Any]], reviewer_dir: Path) -> None:
    wb = Workbook()
    ws_intro = wb.active
    ws_intro.title = "Instructions"

    intro_lines = [
        "Claim 复核说明",
        "1. 先阅读 CLAIM_REVIEW_GUIDE.md，再开始填写。",
        "2. 打开 Answers 工作表，每行点击“查看卡片”。",
        "3. 只填写实体诉求；程序性费用、上诉外壳、证据目录不要写。",
        "4. 必须从原文精确复制，不要改写。",
        "5. 一句中若存在两个可独立裁判的救济效果，应拆成两条填写。",
        "6. 若该案没有实体诉求，claim_1 至 claim_8 留空，并在 notes 中写“无实体诉求”。",
        "7. 例：`判令被告支付货款10000元，并承担违约金3000元` 应拆为两条。",
        "8. 例：`本案诉讼费由被告承担` 不纳入 Claim。",
    ]

    for idx, text in enumerate(intro_lines, start=1):
        ws_intro[f"A{idx}"] = text

    ws_intro["A1"].font = Font(bold=True, size=14)
    ws_intro.column_dimensions["A"].width = 110

    for row in ws_intro.iter_rows(
        min_row=1, max_row=len(intro_lines), min_col=1, max_col=1
    ):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    ws = wb.create_sheet("Answers")

    headers = [
        "anchor_case_id",
        "stage",
        "查看卡片",
        "claim_1",
        "claim_2",
        "claim_3",
        "claim_4",
        "claim_5",
        "claim_6",
        "claim_7",
        "claim_8",
        "notes",
        "_uid",
    ]

    ws.append(headers)
    _apply_header_style(ws)
    ws.freeze_panes = "A2"

    widths = {
        "A": 18,
        "B": 10,
        "C": 14,
        "D": 28,
        "E": 28,
        "F": 28,
        "G": 28,
        "H": 28,
        "I": 28,
        "J": 28,
        "K": 28,
        "L": 28,
        "M": 18,
    }

    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    ws.column_dimensions["M"].hidden = True

    for row_idx, row in enumerate(rows, start=2):
        anchor = str(row["anchor_case_id"])
        card_name = f"{_slug(anchor)}.html"
        ws.cell(row=row_idx, column=1, value=anchor)
        ws.cell(row=row_idx, column=2, value=str(row.get("stage", "") or ""))
        link_cell = ws.cell(row=row_idx, column=3, value="打开卡片")
        link_cell.hyperlink = f"claim_cards/{card_name}"
        link_cell.style = "Hyperlink"
        ws.cell(row=row_idx, column=13, value=str(row.get("uid", "") or ""))

        for col in range(4, 13):
            ws.cell(row=row_idx, column=col).alignment = Alignment(
                vertical="top", wrap_text=True
            )

    wb.save(reviewer_dir / "claim_review.xlsx")


def _build_status_workbook(rows: list[dict[str, Any]], reviewer_dir: Path) -> None:
    wb = Workbook()
    ws_intro = wb.active
    ws_intro.title = "Instructions"

    intro_lines = [
        "Status 复核说明",
        "1. 先阅读 STATUS_REVIEW_GUIDE.md，再开始填写。",
        "2. 打开 Answers 工作表，每行点击“查看卡片”。",
        "3. 先看待判定 Claim，再看裁判结果；不够清楚时再看法院观点。",
        "4. 只在 status 列选择 ACCEPTED / REJECTED / UNMENTIONED。",
        "5. 不要修改 claim_text_raw。",
        "6. 仅部分支持、金额调减、范围缩窄、只支持子项统一记为 UNMENTIONED。",
        "7. 例：判决支持10000元中的6000元，应选 UNMENTIONED，不是 ACCEPTED。",
        "8. 例：二审“驳回上诉，维持原判”，且该 Claim 明显属于上诉人的实体请求，应选 REJECTED。",
    ]

    for idx, text in enumerate(intro_lines, start=1):
        ws_intro[f"A{idx}"] = text

    ws_intro["A1"].font = Font(bold=True, size=14)
    ws_intro.column_dimensions["A"].width = 110

    for row in ws_intro.iter_rows(
        min_row=1, max_row=len(intro_lines), min_col=1, max_col=1
    ):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    ws = wb.create_sheet("Answers")

    headers = [
        "anchor_case_id",
        "stage",
        "claim_text_raw",
        "查看卡片",
        "status",
        "notes",
        "_uid",
        "_claim_id",
    ]

    ws.append(headers)
    _apply_header_style(ws)
    ws.freeze_panes = "A2"
    widths = {"A": 22, "B": 10, "C": 48, "D": 14, "E": 18, "F": 28, "G": 18, "H": 26}

    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    ws.column_dimensions["G"].hidden = True
    ws.column_dimensions["H"].hidden = True

    validation = DataValidation(
        type="list", formula1='"ACCEPTED,REJECTED,UNMENTIONED"', allow_blank=True
    )

    ws.add_data_validation(validation)

    for row_idx, row in enumerate(rows, start=2):
        anchor = str(row["anchor_case_id"])
        card_name = f"{_slug(anchor)}.html"

        claim_text = str(
            (row.get("claims_review") or [{}])[0].get("claim_text_raw", "") or ""
        )

        ws.cell(row=row_idx, column=1, value=anchor)
        ws.cell(row=row_idx, column=2, value=str(row.get("stage", "") or ""))
        claim_cell = ws.cell(row=row_idx, column=3, value=claim_text)
        claim_cell.alignment = Alignment(vertical="top", wrap_text=True)
        link_cell = ws.cell(row=row_idx, column=4, value="打开卡片")
        link_cell.hyperlink = f"status_cards/{card_name}"
        link_cell.style = "Hyperlink"
        validation.add(ws.cell(row=row_idx, column=5))

        ws.cell(row=row_idx, column=6).alignment = Alignment(
            vertical="top", wrap_text=True
        )

        ws.cell(row=row_idx, column=7, value=str(row.get("uid", "") or ""))
        ws.cell(row=row_idx, column=8, value=str(row.get("claim_id", "") or ""))

    wb.save(reviewer_dir / "status_review.xlsx")


def _build_cards(rows: list[dict[str, Any]], card_dir: Path, task: str) -> None:
    card_dir.mkdir(parents=True, exist_ok=True)
    names = [f"{_slug(str(row['anchor_case_id']))}.html" for row in rows]

    for idx, row in enumerate(rows):
        prev_name = names[idx - 1] if idx > 0 else None
        next_name = names[idx + 1] if idx + 1 < len(rows) else None

        content = (
            _render_claim_card(row, prev_name, next_name)
            if task == "claim"
            else _render_status_card(row, prev_name, next_name)
        )

        _write_text(card_dir / names[idx], content)

    title = "Claim 复核目录" if task == "claim" else "Status 复核目录"
    subtitle = "请从此处打开卡片，或使用 Excel 中的超链接。"

    _write_text(
        card_dir / "index.html",
        _render_index(title, subtitle, rows, card_dir.name, task),
    )


def _write_reviewer_readme(
    reviewer_dir: Path, reviewer_label: str, claim_count: int, status_count: int
) -> None:
    text = f"""# {reviewer_label} 人审包

## 你需要填写的文件

- `claim_review.xlsx`
- `status_review.xlsx`

## 开始前必须先读

- `CLAIM_REVIEW_GUIDE.md`
- `STATUS_REVIEW_GUIDE.md`

## 阅读材料

- `claim_cards/index.html`
- `status_cards/index.html`

## 工作顺序

1. 打开 `claim_review.xlsx`，按行点击“打开卡片”，阅读 `claim_cards/` 中的内容并填写答案。
2. 打开 `status_review.xlsx`，按行点击“打开卡片”，阅读 `status_cards/` 中的内容并填写答案。
3. 不要修改 HTML 卡片，不要改文件名。
4. 填写完成后，直接返回这两个 Excel 文件即可。

## 当前样本量

- Claim: {claim_count}
- Status: {status_count}

## 关键规则

- Claim: 必须从原文精确复制，不要改写。
- Status: 只填 `ACCEPTED / REJECTED / UNMENTIONED`。
- 如果拿不准，请先按你理解填写，并把理由写到 `notes`。
"""
    _write_text(reviewer_dir / "START_HERE.md", text)


def _build_reviewer_package(
    reviewer_dir: Path,
    reviewer_label: str,
    claim_rows: list[dict[str, Any]],
    status_rows: list[dict[str, Any]],
) -> None:
    reviewer_dir.mkdir(parents=True, exist_ok=True)

    _write_reviewer_readme(
        reviewer_dir, reviewer_label, len(claim_rows), len(status_rows)
    )

    _write_text(reviewer_dir / "CLAIM_REVIEW_GUIDE.md", _claim_guide_text())
    _write_text(reviewer_dir / "STATUS_REVIEW_GUIDE.md", _status_guide_text())
    _build_cards(claim_rows, reviewer_dir / "claim_cards", "claim")
    _build_cards(status_rows, reviewer_dir / "status_cards", "status")
    _build_claim_workbook(claim_rows, reviewer_dir)
    _build_status_workbook(status_rows, reviewer_dir)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Build a non-technical human review delivery package."
    )

    parser.add_argument(
        "--claim-blind-a",
        default="benchmarks/experiments/artifacts/gold/claim_anchor_review/anchor_set_blind_a.jsonl",
    )

    parser.add_argument(
        "--claim-blind-b",
        default="benchmarks/experiments/artifacts/gold/claim_anchor_review/anchor_set_blind_b.jsonl",
    )

    parser.add_argument(
        "--status-blind-a",
        default="benchmarks/experiments/artifacts/gold/status_anchor_review/status_anchor_set_blind_a.jsonl",
    )

    parser.add_argument(
        "--status-blind-b",
        default="benchmarks/experiments/artifacts/gold/status_anchor_review/status_anchor_set_blind_b.jsonl",
    )

    parser.add_argument("--output-dir", required=True)
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    claim_a = _load_jsonl(Path(args.claim_blind_a))
    claim_b = _load_jsonl(Path(args.claim_blind_b))
    status_a = _load_jsonl(Path(args.status_blind_a))
    status_b = _load_jsonl(Path(args.status_blind_b))
    out = Path(args.output_dir)
    out.mkdir(parents=True, exist_ok=True)
    top_readme = f"""# 人审交付包

本目录可直接打包发送。

## 目录结构

- `reviewer_a/`：标注人 A 使用
- `reviewer_b/`：标注人 B 使用

每位标注人目录内包含：

- `claim_review.xlsx`
- `status_review.xlsx`
- `claim_cards/`
- `status_cards/`
- `START_HERE.md`
- `CLAIM_REVIEW_GUIDE.md`
- `STATUS_REVIEW_GUIDE.md`

## 样本量

- Claim: {len(claim_a)} / {len(claim_b)}
- Status: {len(status_a)} / {len(status_b)}

建议将 `reviewer_a/` 与 `reviewer_b/` 分别单独打包发送，并要求标注人在填写前先阅读各自目录中的两份 Guide。
"""

    _write_text(out / "README.md", top_readme)

    manifest = {
        "claim_count_reviewer_a": len(claim_a),
        "claim_count_reviewer_b": len(claim_b),
        "status_count_reviewer_a": len(status_a),
        "status_count_reviewer_b": len(status_b),
    }

    _write_text(
        out / "package_manifest.json",
        json.dumps(manifest, ensure_ascii=False, indent=2),
    )

    _build_reviewer_package(out / "reviewer_a", "标注人 A", claim_a, status_a)
    _build_reviewer_package(out / "reviewer_b", "标注人 B", claim_b, status_b)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
